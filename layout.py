from openpyxl import load_workbook, Workbook
from random import seed
from random import randint
import companies as c
seed(1)
class Booth(object):
    def __init__(self,boothName, boothRow, boothCol, previousBooth=None,companyName="",isPower=False, isPremium = False):
        self.boothName = boothName
        self.company_name = "" #could replace this with company data type
        self.boothRow = boothRow
        self.boothCol = boothCol
        self.previousBooth= previousBooth
        self.isPower = isPower
        self.isPremium = isPremium
        if(self.isPremium):
            self.isPower = True
    def makePremium(self):
        self.isPremium = True
        self.isPower = True
    def show(self):
        print()
        print(str(self.boothName))
        print("Row :"+ str(self.boothRow))
        print("Col :" + str(self.boothCol))
        print("Is Power :" + str(self.isPower))
        print("Company :" + self.company_name)
            


def makeBooths(startrow, endrow, startcol, endcol):
    wb = load_workbook(filename = 'CRC floor diagram 8x16 v2.xlsx')
    sheet_ranges = wb['Courts 3-6 only']
    boothArr = []

    for col in range(startcol, endcol+1):
        for row in range(startrow, endrow + 1):
            # print(sheet_ranges.cell(column=col, row=row).value)
            cell = sheet_ranges.cell(column=col, row=row)
            # print(type(cell).__name__)
            if(str(cell.value) != 'None'):
                booth = Booth(cell.value, row, col)
                if '*' in booth.boothName:
                    booth.makePremium()
                elif cell.fill.start_color.rgb == 'FFFFFF00':
                    booth.isPower = True
                boothArr = boothArr + [booth]
    return boothArr


startrow = 26
endrow = 73
startcol = 4
endcol = 30
boothArr = makeBooths(startrow, endrow, startcol, endcol)
boothArr = boothArr + makeBooths(28,71,2,3)
boothArr = boothArr + makeBooths(28,69,32,33)
boothHash= {}
for b in boothArr:
    if not (b.boothName[1:] in boothHash):
        if b.boothName[-1] == '*':
            if not((b.boothName[:-1])[1:] in boothHash):
                # Add booth number to the hash
                boothHash[(b.boothName[:-1])[1:]] = [b.boothName[0]]
            else:
                boothHash[(b.boothName[:-1])[1:]] = boothHash[(b.boothName[:-1])[1:]] + [b.boothName[0]]
        else:
            boothHash[b.boothName[1:]] = [b.boothName[0]]
    else:
        boothHash[b.boothName[1:]] = boothHash[b.boothName[1:]] + [b.boothName[0]]

def filterPremiumBooths(boothAr):
    premBooths = []
    for booth in boothAr:
        if '*' in booth.boothName:
            premBooths = premBooths + [booth]
    return premBooths
premComps = filterPremiumBooths(boothArr)

def filterPowerBooths(boothAr):
    powBooths = []
    for booth in boothAr:
        if booth.isPower:
            powBooths = powBooths + [booth]
    return powBooths
powComps = filterPowerBooths(boothArr)

boothArr.sort(key=lambda x: x.boothName)

finBooths = []
comps = c.getCompanies()
standardBoothIndex = 0
for comp in comps:
    if 'Premium Booth' in comp.booths:
       print(comp.booths)
       pBooth =  premComps[-1]
       pBooth.company_name = comp.employer
       finBooths = finBooths + [pBooth]
       premComps = premComps[0:-1]
    elif comp.needsElectric:
        powBooth = powComps[-1]
        powBooth.company_name = comp.employer
        finBooths = finBooths + [powBooth]
        powComps = powComps[0:-1]
    else:
        sBooth = None
        while sBooth == None:
            if (not boothArr[standardBoothIndex].isPremium) and (not boothArr[standardBoothIndex].isPower):
               if boothArr[standardBoothIndex].company_name == '':
                   boothArr[standardBoothIndex].company_name = comp.employer
                   finBooths = finBooths + [boothArr[standardBoothIndex]]
                   sBooth = boothArr[standardBoothIndex]
            standardBoothIndex = standardBoothIndex + 1
            
        # Add standard booths to finbooth Array
        # Try to make each industry assigned to a letter booth
        # ex- civils to A, EECS to B, etc
    

# Assigning the booths to the excel file
wb = load_workbook(filename = 'CRC floor diagram 8x16 v2 test.xlsx')
sheet_ranges = wb['Courts 3-6 only']

for b in finBooths:
    cell = sheet_ranges.cell(column=b.boothCol, row=b.boothRow)
    cell.value = b.company_name
wb.save('CRC floor diagram 8x16 v2 Test1.xlsx')
print(boothHash)






            
            



# sheet_ranges.unmerge_cells('C5:D6')




